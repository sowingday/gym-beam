from __future__ import annotations

import hashlib
import json
import re
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


COLUMN_ALIASES = {
    "index": ["index", "idx", "nr", "nummer"],
    "name": ["name", "uebung", "übung", "exercise", "exercise name", "exercisename"],
    "nameEn": ["name en", "name (en)", "name englisch", "exercise name en", "exercise name (en)"],
    "gender": ["gender", "geschlecht", "sex"],
    "genderVariantGroup": ["gender variant group", "gender_variant_group", "gendergruppe", "gender gruppe", "geschlechtsvariante gruppe", "geschlechtsvarianten gruppe"],
    "categories": ["kategorien", "kategorie", "categories", "category"],
    "categoriesEn": ["kategorien en", "kategorien (en)", "categories en", "categories (en)", "category en"],
    "muscles": ["muskeln", "muskel", "muscles", "muscle"],
    "musclesEn": ["muskeln en", "muskeln (en)", "muscles en", "muscles (en)", "muscle en"],
    "musclesLatin": ["muskeln lat", "muskeln lat.", "muskeln (lat.)", "muskeln latein", "muskeln latin", "muscles latin", "latin muscles"],
    "shortDescription": ["kurzbeschreibung", "beschreibung", "shortdescription", "short description", "description"],
    "shortDescriptionEn": ["kurzbeschreibung en", "kurzbeschreibung (en)", "shortdescription en", "short description en", "short description (en)", "description en"],
    "notes": ["hinweise", "hinweis", "notes", "tips", "remarks"],
    "notesEn": ["hinweise en", "hinweise (en)", "notes en", "notes (en)", "tips en"],
    "author": ["autor", "author"],
    "packName": ["iconscout-pack", "pack", "pack name", "iconscout pack"],
    "packId": ["pack-id", "pack id", "packid"],
    "originalName": ["original name", "originalname"],
}

REQUIRED_FIELDS = ["index", "name"]
OPTIONAL_FIELDS = [
    "nameEn",
    "gender",
    "genderVariantGroup",
    "categories",
    "categoriesEn",
    "muscles",
    "musclesEn",
    "musclesLatin",
    "shortDescription",
    "shortDescriptionEn",
    "notes",
    "notesEn",
    "author",
    "packName",
    "packId",
    "originalName",
]

EXPORT_FIELDS = [
    "index",
    "name",
    "nameEn",
    "gender",
    "genderVariantGroup",
    "categories",
    "categoriesEn",
    "muscles",
    "musclesEn",
    "musclesLatin",
    "shortDescription",
    "shortDescriptionEn",
    "notes",
    "notesEn",
]

FIELD_LABELS = {
    "name": "Name (DE)",
    "nameEn": "Name (EN)",
    "gender": "Gender (m/w)",
    "genderVariantGroup": "Gender-Variante Gruppe",
    "categories": "Kategorien (DE)",
    "categoriesEn": "Kategorien (EN)",
    "muscles": "Muskeln (DE)",
    "musclesEn": "Muskeln (EN)",
    "musclesLatin": "Muskeln (lat.)",
    "shortDescription": "Kurzbeschreibung (DE)",
    "shortDescriptionEn": "Kurzbeschreibung (EN)",
    "notes": "Hinweise (DE)",
    "notesEn": "Hinweise (EN)",
}

AI_SUGGEST_FIELDS = [
    "name",
    "nameEn",
    "gender",
    "genderVariantGroup",
    "categories",
    "categoriesEn",
    "muscles",
    "musclesEn",
    "musclesLatin",
    "shortDescription",
    "shortDescriptionEn",
    "notes",
    "notesEn",
]

SUSPICIOUS_PATTERNS = [
    r"\bman\b",
    r"\bwoman\b",
    r"\bdoing\b",
    r"\bexercise for\b",
    r"\bin place\b",
    r"--",
]

BODYWEIGHT_LABELS = {
    "ohne gerät",
    "ohne geraet",
    "bodyweight",
}

EQUIPMENT_HINT_PATTERNS = [
    r"\bab rollout\b",
    r"\bwheel\b",
    r"\broller\b",
    r"\btrx\b",
    r"\bsuspension\b",
    r"\bdumbbell\b",
    r"\bbarbell\b",
    r"\bkettlebell\b",
    r"\bcable\b",
    r"\bmachine\b",
    r"\bsmith\b",
    r"\bbench\b",
    r"\bband\b",
    r"\bresistance band\b",
    r"\bplate\b",
    r"\bball\b",
    r"\bbosu\b",
    r"\bstep box\b",
    r"\blandmine\b",
    r"\bcaptain'?s chair\b",
    r"\bpull up\b",
    r"\bdip\b",
]


def norm_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    text = re.sub(r"[()\[\]{}.:;,_\-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


ALIAS_LOOKUP = {
    field_name: {norm_header(alias) for alias in aliases}
    for field_name, aliases in COLUMN_ALIASES.items()
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def split_list(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    parts = re.split(r"[;,\n]+", text)
    return [part.strip() for part in parts if part and part.strip()]


def join_list(values: list[str]) -> str:
    return "; ".join(value.strip() for value in values if value and value.strip())


def toggle_list_value(current_text: str, current_text_en: str, de_value: str, en_value: str, checked: bool) -> tuple[str, str]:
    current_de = [item.strip() for item in split_list(current_text) if item.strip()]
    current_en = [item.strip() for item in split_list(current_text_en) if item.strip()]
    lowered = [item.casefold() for item in current_de]
    if checked:
        if de_value.casefold() not in lowered:
            current_de.append(de_value)
            if en_value:
                current_en.append(en_value)
    else:
        filtered_pairs = [
            (de_item, current_en[idx] if idx < len(current_en) else "")
            for idx, de_item in enumerate(current_de)
            if de_item.casefold() != de_value.casefold()
        ]
        current_de = [pair[0] for pair in filtered_pairs]
        current_en = [pair[1] for pair in filtered_pairs if pair[1]]
    return join_list(current_de), join_list(current_en)


def parse_index(value: object) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        match = re.search(r"\d+", text)
        return int(match.group(0)) if match else None


def normalize_name_key(value: str) -> str:
    return re.sub(r"\s+", " ", clean_text(value).casefold())


def normalize_gender_value(value: object) -> str:
    text = clean_text(value).casefold()
    if text in {"m", "male", "mann", "maennlich", "männlich"}:
        return "m"
    if text in {"w", "female", "frau", "weiblich"}:
        return "w"
    return ""


def normalize_variant_group(value: object) -> str:
    return clean_text(value)


def make_i18n(de_value: object, en_value: object) -> dict[str, Any]:
    de = clean_text(de_value)
    en = clean_text(en_value)
    payload: dict[str, Any] = {"de": de}
    if en:
        payload["en"] = en
    return payload


def make_i18n_list(de_value: object, en_value: object) -> dict[str, Any]:
    de = split_list(de_value)
    en = split_list(en_value)
    payload: dict[str, Any] = {"de": de}
    if en:
        payload["en"] = en
    return payload


def looks_suspicious_name(text: str) -> bool:
    value = clean_text(text)
    if not value:
        return True
    lowered = value.casefold()
    return any(re.search(pattern, lowered) for pattern in SUSPICIOUS_PATTERNS)


def slug_to_title(text: str) -> str:
    value = clean_text(text)
    value = re.sub(r"[_/]+", " ", value)
    value = re.sub(r"[-]+", " ", value)
    value = re.sub(r"\b(man|woman|doing|exercise|for|workout|animated|animation)\b", " ", value, flags=re.I)
    value = re.sub(r"\s+", " ", value).strip(" -")
    return value.title()


def local_name_suggestion(entry: "ExerciseEntry") -> dict[str, str]:
    candidates = [
        entry.values.get("nameEn", ""),
        entry.values.get("originalName", ""),
        entry.values.get("name", ""),
    ]
    best_en = ""
    for candidate in candidates:
        cleaned = slug_to_title(candidate)
        if cleaned and len(cleaned) >= 3:
            best_en = cleaned
            break

    best_de = clean_text(entry.values.get("name", ""))
    if not best_de or looks_suspicious_name(best_de):
        equipment_prefix = ""
        first_category = split_list(entry.values.get("categories", ""))
        if first_category:
            normalized = first_category[0].casefold()
            equipment_map = {
                "kurzhantel": "Kurzhantel-",
                "langhantel": "Langhantel-",
                "widerstandsband": "Widerstandsband-",
                "trx schlingentrainer": "TRX-",
                "gymnastikball": "Gymnastikball-",
            }
            equipment_prefix = equipment_map.get(normalized, "")
        best_de = f"{equipment_prefix}{best_en}".strip("- ") if best_en else best_de

    reason_parts = []
    if entry.values.get("originalName"):
        reason_parts.append(f"Originalname: {entry.values['originalName']}")
    if entry.values.get("nameEn"):
        reason_parts.append(f"Aktueller EN-Name: {entry.values['nameEn']}")
    if looks_suspicious_name(entry.values.get("name", "")):
        reason_parts.append("Aktueller Name wirkt automatisch erzeugt oder fehlerhaft.")

    return {
        "name": best_de,
        "nameEn": best_en,
        "reason": " | ".join(reason_parts) if reason_parts else "Lokaler Heuristik-Vorschlag.",
    }


def local_full_suggestion(entry: "ExerciseEntry") -> dict[str, str]:
    base = local_name_suggestion(entry)
    payload = {
        field_name: clean_text(entry.values.get(field_name, ""))
        for field_name in AI_SUGGEST_FIELDS
    }
    payload["name"] = clean_text(base.get("name", payload.get("name", "")))
    payload["nameEn"] = clean_text(base.get("nameEn", payload.get("nameEn", "")))
    payload["reason"] = clean_text(base.get("reason", "Lokaler Heuristik-Vorschlag."))
    return payload


def extract_lottie_name_hints(animation_path: Path) -> list[str]:
    if not animation_path.exists():
        return []
    try:
        data = json.loads(animation_path.read_text(encoding="utf-8"))
    except Exception:
        return []

    hints: list[str] = []

    def walk(node: Any) -> None:
        if isinstance(node, dict):
            value = node.get("nm")
            if isinstance(value, str):
                cleaned = clean_text(value)
                if cleaned and cleaned not in hints:
                    hints.append(cleaned)
            for child in node.values():
                if len(hints) >= 24:
                    return
                walk(child)
        elif isinstance(node, list):
            for child in node:
                if len(hints) >= 24:
                    return
                walk(child)

    walk(data)
    return hints[:24]


def ai_name_suggestion(entry: "ExerciseEntry", config: AppConfig, animation_path: Path) -> dict[str, str]:
    if not config.ai_enabled:
        raise RuntimeError("KI-Anbieter ist in den Einstellungen nicht aktiviert.")
    if not config.ai_api_key.strip():
        raise RuntimeError("Kein API-Key für den KI-Anbieter hinterlegt.")

    prompt = {
        "task": "Normalize and improve exercise naming and classification for a fitness app.",
        "language": "Return concise German and English names plus optional improved categories and muscles.",
        "current": {
            "index": entry.index,
            "name": entry.values.get("name", ""),
            "nameEn": entry.values.get("nameEn", ""),
            "categories": split_list(entry.values.get("categories", "")),
            "categoriesEn": split_list(entry.values.get("categoriesEn", "")),
            "muscles": split_list(entry.values.get("muscles", "")),
            "musclesEn": split_list(entry.values.get("musclesEn", "")),
            "originalName": entry.values.get("originalName", ""),
            "packName": entry.values.get("packName", ""),
        },
        "animationHints": extract_lottie_name_hints(animation_path),
        "requirements": [
            "Return strict JSON only.",
            "Use keys: name, nameEn, categories, muscles, short_reason.",
            "categories and muscles must be arrays of German labels.",
            "Prefer established fitness terminology.",
            "If uncertain, keep close to original movement."
        ],
    }

    body = {
        "model": config.ai_model,
        "messages": [
            {"role": "system", "content": "You are a fitness exercise taxonomy assistant. Reply with strict JSON only."},
            {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
        ],
        "temperature": 0.2,
        "response_format": {"type": "json_object"},
    }

    base_url = config.ai_base_url.rstrip("/")
    request = urllib.request.Request(
        f"{base_url}/chat/completions",
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {config.ai_api_key}",
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=45) as response:
        payload = json.loads(response.read().decode("utf-8"))

    content = payload["choices"][0]["message"]["content"]
    data = json.loads(content)
    return {
        "name": clean_text(data.get("name", "")),
        "nameEn": clean_text(data.get("nameEn", "")),
        "categories": join_list(data.get("categories", []) if isinstance(data.get("categories"), list) else split_list(data.get("categories", ""))),
        "muscles": join_list(data.get("muscles", []) if isinstance(data.get("muscles"), list) else split_list(data.get("muscles", ""))),
        "reason": clean_text(data.get("short_reason", "")),
    }


def _normalize_ai_field(field_name: str, value: Any) -> str:
    if field_name in {"categories", "categoriesEn", "muscles", "musclesEn", "musclesLatin"}:
        if isinstance(value, list):
            return join_list([clean_text(item) for item in value])
        return join_list(split_list(value))
    if field_name == "gender":
        return normalize_gender_value(value)
    return clean_text(value)


def ai_bulk_suggestions(entries: list["ExerciseEntry"], config: AppConfig, animation_paths: dict[int, Path]) -> dict[int, dict[str, str]]:
    if not entries:
        return {}
    if not config.ai_enabled:
        raise RuntimeError("KI-Anbieter ist in den Einstellungen nicht aktiviert.")
    if not config.ai_api_key.strip():
        raise RuntimeError("Kein API-Key für den KI-Anbieter hinterlegt.")

    prompt_items: list[dict[str, Any]] = []
    for entry in entries:
        prompt_items.append(
            {
                "index": entry.index,
                "current": {
                    field_name: entry.values.get(field_name, "")
                    for field_name in [*AI_SUGGEST_FIELDS, "originalName", "packName"]
                },
                "animationHints": extract_lottie_name_hints(animation_paths.get(entry.index, Path()))[:12],
            }
        )

    prompt = {
        "task": "Normalize and improve exercise metadata for a fitness app.",
        "language": "Return concise German and English labels and text. Use strict JSON only.",
        "items": prompt_items,
        "requirements": [
            "Return JSON with top-level key 'items'.",
            "Each item must contain: index, name, nameEn, gender, genderVariantGroup, categories, categoriesEn, muscles, musclesEn, musclesLatin, shortDescription, shortDescriptionEn, notes, notesEn, short_reason.",
            "categories, categoriesEn, muscles, musclesEn, musclesLatin must be arrays.",
            "gender must be one of: m, w, or empty string.",
            "Only set genderVariantGroup if the pairing is very clear; otherwise return empty string.",
            "Keep close to the original exercise if uncertain.",
        ],
    }

    body = {
        "model": config.ai_model,
        "messages": [
            {"role": "system", "content": "You are a fitness exercise taxonomy assistant. Reply with strict JSON only."},
            {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
        ],
        "temperature": 0.2,
        "response_format": {"type": "json_object"},
    }

    base_url = config.ai_base_url.rstrip("/")
    request = urllib.request.Request(
        f"{base_url}/chat/completions",
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {config.ai_api_key}",
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=90) as response:
        payload = json.loads(response.read().decode("utf-8"))

    content = payload["choices"][0]["message"]["content"]
    data = json.loads(content)
    result: dict[int, dict[str, str]] = {}
    for item in data.get("items", []):
        try:
            index_value = int(item.get("index"))
        except Exception:
            continue
        normalized = {
            field_name: _normalize_ai_field(field_name, item.get(field_name, ""))
            for field_name in AI_SUGGEST_FIELDS
        }
        normalized["reason"] = clean_text(item.get("short_reason", ""))
        result[index_value] = normalized
    return result


@dataclass
class AppConfig:
    excel_file: str
    app_json_file: str
    animations_dir: str
    export_report_file: str
    ai_enabled: bool
    ai_base_url: str
    ai_model: str
    ai_api_key: str

    @classmethod
    def from_dict(cls, payload: dict[str, Any]) -> "AppConfig":
        return cls(
            excel_file=str(payload.get("excel_file", "")),
            app_json_file=str(payload.get("app_json_file", "")),
            animations_dir=str(payload.get("animations_dir", "")),
            export_report_file=str(payload.get("export_report_file", "")),
            ai_enabled=bool(payload.get("ai_enabled", False)),
            ai_base_url=str(payload.get("ai_base_url", "https://api.openai.com/v1")),
            ai_model=str(payload.get("ai_model", "gpt-4.1-mini")),
            ai_api_key=str(payload.get("ai_api_key", "")),
        )

    def to_dict(self) -> dict[str, str]:
        return {
            "excel_file": self.excel_file,
            "app_json_file": self.app_json_file,
            "animations_dir": self.animations_dir,
            "export_report_file": self.export_report_file,
            "ai_enabled": self.ai_enabled,
            "ai_base_url": self.ai_base_url,
            "ai_model": self.ai_model,
            "ai_api_key": self.ai_api_key,
        }


class ConfigStore:
    def __init__(self, base_dir: Path):
        self.base_dir = Path(base_dir)
        self.defaults_path = self.base_dir / "exercise_audit_tool.defaults.json"
        self.local_path = self.base_dir / "exercise_audit_tool.local.json"

    def load(self) -> AppConfig:
        payload: dict[str, Any] = {}
        if self.defaults_path.exists():
            payload.update(json.loads(self.defaults_path.read_text(encoding="utf-8")))
        if self.local_path.exists():
            payload.update(json.loads(self.local_path.read_text(encoding="utf-8")))
        return AppConfig.from_dict(payload)

    def save_local(self, config: AppConfig) -> None:
        self.local_path.write_text(json.dumps(config.to_dict(), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


@dataclass
class ExerciseEntry:
    row_index: int
    index: int
    values: dict[str, str] = field(default_factory=dict)
    duplicate_name_count: int = 1
    suspicious_name: bool = False
    animation_exists: bool = False
    json_mismatch: bool = False
    json_mismatch_fields: list[str] = field(default_factory=list)
    equipment_conflict: bool = False
    equipment_conflict_reason: str = ""

    @property
    def display_name(self) -> str:
        return self.values.get("name", "")

    @property
    def status_flags(self) -> list[str]:
        flags: list[str] = []
        if not self.animation_exists:
            flags.append("Animation fehlt")
        if self.json_mismatch:
            flags.append("JSON abweichend")
        if self.equipment_conflict:
            flags.append("Kategorie prüfen")
        if self.suspicious_name:
            flags.append("Verdächtiger Name")
        if self.duplicate_name_count > 1:
            flags.append("Duplikatname")
        if not flags:
            flags.append("OK")
        return flags

    @property
    def status_text(self) -> str:
        return " | ".join(self.status_flags)


class WorkbookAuditStore:
    def __init__(self, excel_path: Path, animations_dir: Path | None = None):
        self.excel_path = Path(excel_path)
        self.animations_dir = Path(animations_dir) if animations_dir else None
        self.sheet_name = ""
        self.header_row = 0
        self.column_map: dict[str, int] = {}
        self.workbook = None
        self.sheet = None
        self.entries: list[ExerciseEntry] = []
        self.entries_by_index: dict[int, ExerciseEntry] = {}
        self.app_json_snapshot: dict[int, dict[str, Any]] = {}

    def load(self) -> None:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel-Datei nicht gefunden: {self.excel_path}")

        workbook = load_workbook(self.excel_path)
        best_score = -1.0
        best_sheet = None
        best_header_row = None
        best_column_map = None

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_scan = min(sheet.max_row, 80)
            for row_number in range(1, max_scan + 1):
                found: dict[str, int] = {}
                for col_number in range(1, sheet.max_column + 1):
                    value = norm_header(sheet.cell(row=row_number, column=col_number).value)
                    if not value:
                        continue
                    for field_name, aliases in ALIAS_LOOKUP.items():
                        if value in aliases and field_name not in found:
                            found[field_name] = col_number
                required_found = [field_name for field_name in REQUIRED_FIELDS if field_name in found]
                optional_found = [field_name for field_name in OPTIONAL_FIELDS if field_name in found]
                score = len(required_found) + 0.25 * len(optional_found)
                if "index" in found and "name" in found and score > best_score:
                    best_score = score
                    best_sheet = sheet_name
                    best_header_row = row_number
                    best_column_map = found

        if not best_sheet or not best_column_map or best_header_row is None:
            raise RuntimeError("Keine passende Kopfzeile in der Excel-Datei gefunden.")

        self.workbook = workbook
        self.sheet_name = best_sheet
        self.header_row = best_header_row
        self.column_map = best_column_map
        self.sheet = workbook[best_sheet]
        self._load_entries()

    def _load_entries(self) -> None:
        assert self.sheet is not None
        grouped_by_name: dict[str, list[ExerciseEntry]] = {}
        entries: list[ExerciseEntry] = []

        for row_index in range(self.header_row + 1, self.sheet.max_row + 1):
            raw_index = self._get_cell_value(row_index, "index")
            raw_name = self._get_cell_value(row_index, "name")
            index_value = parse_index(raw_index)
            name_value = clean_text(raw_name)
            if index_value is None and not name_value:
                continue
            if index_value is None:
                continue

            values = {field_name: clean_text(self._get_cell_value(row_index, field_name)) for field_name in self.column_map.keys()}
            entry = ExerciseEntry(
                row_index=row_index,
                index=index_value,
                values=values,
                suspicious_name=looks_suspicious_name(values.get("name", "")) or looks_suspicious_name(values.get("nameEn", "")),
                animation_exists=self.animation_path_for(index_value).exists() if self.animations_dir else False,
            )
            entry.equipment_conflict, entry.equipment_conflict_reason = self._detect_equipment_conflict(entry)
            entries.append(entry)
            key = normalize_name_key(values.get("name", ""))
            if key:
                grouped_by_name.setdefault(key, []).append(entry)

        for group_entries in grouped_by_name.values():
            duplicate_count = self._unresolved_duplicate_count(group_entries)
            for entry in group_entries:
                entry.duplicate_name_count = duplicate_count

        self.entries = sorted(entries, key=lambda item: item.index)
        self.entries_by_index = {entry.index: entry for entry in self.entries}

    @staticmethod
    def _unresolved_duplicate_count(entries: list[ExerciseEntry]) -> int:
        collapsed_keys: set[str] = set()
        for entry in entries:
            group = normalize_variant_group(entry.values.get("genderVariantGroup", ""))
            if group:
                collapsed_keys.add(f"group:{group.casefold()}")
            else:
                collapsed_keys.add(f"entry:{entry.index}")
        return len(collapsed_keys)

    def _get_cell_value(self, row_index: int, field_name: str) -> Any:
        column = self.column_map.get(field_name)
        if not column or self.sheet is None:
            return ""
        return self.sheet.cell(row=row_index, column=column).value

    def animation_path_for(self, index_value: int) -> Path:
        if not self.animations_dir:
            return Path("__missing_animation_dir__") / f"{index_value}.json"
        return self.animations_dir / f"{index_value}.json"

    def animation_signature_for(self, index_value: int) -> str:
        path = self.animation_path_for(index_value)
        if not path.exists():
            return ""
        return hashlib.sha1(path.read_bytes()).hexdigest()

    def _detect_equipment_conflict(self, entry: ExerciseEntry) -> tuple[bool, str]:
        categories = [clean_text(item).casefold() for item in split_list(entry.values.get("categories", ""))]
        if not any(category in BODYWEIGHT_LABELS for category in categories):
            return False, ""

        sources = [
            clean_text(entry.values.get("name", "")),
            clean_text(entry.values.get("nameEn", "")),
            clean_text(entry.values.get("originalName", "")),
            clean_text(entry.values.get("packName", "")),
        ]
        for source in sources:
            lowered = source.casefold()
            for pattern in EQUIPMENT_HINT_PATTERNS:
                if re.search(pattern, lowered):
                    return True, f"`ohne Gerät` passt evtl. nicht zu Hinweis aus '{source}'."
        return False, ""

    def save_entry(self, index_value: int, updated_values: dict[str, str]) -> ExerciseEntry:
        if index_value not in self.entries_by_index:
            raise KeyError(f"Übung mit Index {index_value} nicht gefunden.")
        if self.workbook is None or self.sheet is None:
            raise RuntimeError("Arbeitsmappe ist nicht geladen.")

        entry = self.entries_by_index[index_value]
        for field_name, value in updated_values.items():
            if field_name not in self.column_map:
                continue
            self.sheet.cell(row=entry.row_index, column=self.column_map[field_name], value=clean_text(value))
            entry.values[field_name] = clean_text(value)

        self.workbook.save(self.excel_path)
        self._load_entries()
        if self.app_json_snapshot:
            self._apply_json_snapshot()
        return self.entries_by_index[index_value]

    def snapshot_entries(self, indices: list[int], fields: list[str] | None = None) -> dict[int, dict[str, str]]:
        snapshot: dict[int, dict[str, str]] = {}
        selected_fields = fields or list(self.column_map.keys())
        for index_value in indices:
            entry = self.entries_by_index.get(index_value)
            if not entry:
                continue
            snapshot[index_value] = {
                field_name: clean_text(entry.values.get(field_name, ""))
                for field_name in selected_fields
            }
        return snapshot

    def save_entries_bulk(self, updates: dict[int, dict[str, str]]) -> list[int]:
        if self.workbook is None or self.sheet is None:
            raise RuntimeError("Arbeitsmappe ist nicht geladen.")
        changed_indices: list[int] = []
        for index_value, updated_values in updates.items():
            entry = self.entries_by_index.get(index_value)
            if not entry:
                continue
            entry_changed = False
            for field_name, value in updated_values.items():
                if field_name not in self.column_map:
                    continue
                cleaned = clean_text(value)
                if clean_text(entry.values.get(field_name, "")) == cleaned:
                    continue
                self.sheet.cell(row=entry.row_index, column=self.column_map[field_name], value=cleaned)
                entry.values[field_name] = cleaned
                entry_changed = True
            if entry_changed:
                changed_indices.append(index_value)
        if changed_indices:
            self.workbook.save(self.excel_path)
            self._load_entries()
            if self.app_json_snapshot:
                self._apply_json_snapshot()
        return changed_indices

    def delete_entries(self, indices: list[int]) -> list[int]:
        if self.workbook is None or self.sheet is None:
            raise RuntimeError("Arbeitsmappe ist nicht geladen.")
        entries = [self.entries_by_index[index_value] for index_value in indices if index_value in self.entries_by_index]
        deleted = [entry.index for entry in entries]
        for entry in sorted(entries, key=lambda item: item.row_index, reverse=True):
            self.sheet.delete_rows(entry.row_index, 1)
        if deleted:
            self.workbook.save(self.excel_path)
            self._load_entries()
            if self.app_json_snapshot:
                self._apply_json_snapshot()
        return deleted

    def duplicate_group_for(self, index_value: int) -> list[ExerciseEntry]:
        entry = self.entries_by_index.get(index_value)
        if not entry:
            return []
        key = normalize_name_key(entry.values.get("name", ""))
        if not key:
            return []
        return [candidate for candidate in self.entries if normalize_name_key(candidate.values.get("name", "")) == key]

    def gender_variant_candidate_groups(self) -> list[list[ExerciseEntry]]:
        grouped: dict[str, list[ExerciseEntry]] = {}
        for entry in self.entries:
            key = normalize_name_key(entry.values.get("name", ""))
            if not key:
                continue
            grouped.setdefault(key, []).append(entry)

        candidates: list[list[ExerciseEntry]] = []
        for group_entries in grouped.values():
            if len(group_entries) < 2:
                continue
            genders = {normalize_gender_value(entry.values.get("gender", "")) for entry in group_entries}
            genders.discard("")
            if len(genders) < 2:
                continue

            variant_groups = {
                normalize_variant_group(entry.values.get("genderVariantGroup", "")).casefold()
                for entry in group_entries
                if normalize_variant_group(entry.values.get("genderVariantGroup", ""))
            }
            # Candidate means there are cross-gender duplicates, but they are not fully resolved
            # into exactly one explicit variant group for the full name group.
            if len(variant_groups) == 1 and all(normalize_variant_group(entry.values.get("genderVariantGroup", "")) for entry in group_entries):
                continue
            candidates.append(sorted(group_entries, key=lambda item: item.index))

        return sorted(candidates, key=lambda group: group[0].index)

    def gender_variant_pair_groups(self) -> list[list[ExerciseEntry]]:
        grouped: dict[tuple[str, str], list[ExerciseEntry]] = {}
        for entry in self.entries:
            name_key = normalize_name_key(entry.values.get("name", ""))
            variant_group = normalize_variant_group(entry.values.get("genderVariantGroup", ""))
            if not name_key or not variant_group:
                continue
            grouped.setdefault((name_key, variant_group.casefold()), []).append(entry)

        pairs: list[list[ExerciseEntry]] = []
        for group_entries in grouped.values():
            if len(group_entries) != 2:
                continue
            genders = [normalize_gender_value(entry.values.get("gender", "")) for entry in group_entries]
            if sorted(genders) != ["m", "w"]:
                continue
            pairs.append(sorted(group_entries, key=lambda item: item.index))
        return sorted(pairs, key=lambda group: group[0].index)

    def available_category_pairs(self) -> list[tuple[str, str]]:
        pairs: dict[str, str] = {}
        for entry in self.entries:
            de_values = split_list(entry.values.get("categories", ""))
            en_values = split_list(entry.values.get("categoriesEn", ""))
            for idx, de_value in enumerate(de_values):
                key = clean_text(de_value)
                if not key:
                    continue
                en_value = en_values[idx] if idx < len(en_values) else ""
                pairs.setdefault(key, clean_text(en_value))
        return sorted(pairs.items(), key=lambda item: item[0].casefold())

    def available_muscle_pairs(self) -> list[tuple[str, str]]:
        pairs: dict[str, str] = {}
        for entry in self.entries:
            de_values = split_list(entry.values.get("muscles", ""))
            en_values = split_list(entry.values.get("musclesEn", ""))
            for idx, de_value in enumerate(de_values):
                key = clean_text(de_value)
                if not key:
                    continue
                en_value = en_values[idx] if idx < len(en_values) else ""
                pairs.setdefault(key, clean_text(en_value))
        return sorted(pairs.items(), key=lambda item: item[0].casefold())

    def load_app_json_snapshot(self, app_json_path: Path | None) -> None:
        self.app_json_snapshot = {}
        if not app_json_path:
            return
        path = Path(app_json_path)
        if not path.exists():
            return
        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            return
        for item in data:
            index_value = parse_index(item.get("index"))
            if index_value is None:
                continue
            self.app_json_snapshot[index_value] = item
        self._apply_json_snapshot()

    def _apply_json_snapshot(self) -> None:
        for entry in self.entries:
            snapshot = self.app_json_snapshot.get(entry.index)
            entry.json_mismatch = False
            entry.json_mismatch_fields = []
            if not snapshot:
                continue
            compare_fields = {
                "name": clean_text(snapshot.get("name", "")),
                "nameEn": clean_text(snapshot.get("nameEn", "")),
                "gender": clean_text(snapshot.get("gender", "")),
                "genderVariantGroup": clean_text(snapshot.get("genderVariantGroup", "")),
                "categories": join_list(snapshot.get("categories", []) if isinstance(snapshot.get("categories"), list) else split_list(snapshot.get("categories", ""))),
                "categoriesEn": join_list(snapshot.get("categoriesEn", []) if isinstance(snapshot.get("categoriesEn"), list) else split_list(snapshot.get("categoriesEn", ""))),
                "muscles": join_list(snapshot.get("muscles", []) if isinstance(snapshot.get("muscles"), list) else split_list(snapshot.get("muscles", ""))),
                "musclesEn": join_list(snapshot.get("musclesEn", []) if isinstance(snapshot.get("musclesEn"), list) else split_list(snapshot.get("musclesEn", ""))),
                "musclesLatin": join_list(snapshot.get("musclesLatin", []) if isinstance(snapshot.get("musclesLatin"), list) else split_list(snapshot.get("musclesLatin", ""))),
                "shortDescription": clean_text(snapshot.get("shortDescription", "")),
                "shortDescriptionEn": clean_text(snapshot.get("shortDescriptionEn", "")),
                "notes": clean_text(snapshot.get("notes", "")),
                "notesEn": clean_text(snapshot.get("notesEn", "")),
            }
            mismatch_fields = [field_name for field_name, json_value in compare_fields.items() if clean_text(entry.values.get(field_name, "")) != json_value]
            entry.json_mismatch = bool(mismatch_fields)
            entry.json_mismatch_fields = mismatch_fields

    def export_json(self, output_json_path: Path, report_path: Path) -> dict[str, Any]:
        output_json_path = Path(output_json_path)
        report_path = Path(report_path)
        exercises: list[dict[str, Any]] = []
        warnings: list[str] = []
        missing_animations: list[str] = []

        for entry in self.entries:
            name = clean_text(entry.values.get("name", ""))
            if not name:
                warnings.append(f"Index {entry.index}: Name fehlt.")
                continue

            item = {
                "index": entry.index,
                "name": name,
                "nameEn": clean_text(entry.values.get("nameEn", "")),
                "gender": normalize_gender_value(entry.values.get("gender", "")),
                "genderVariantGroup": normalize_variant_group(entry.values.get("genderVariantGroup", "")),
                "categories": split_list(entry.values.get("categories", "")),
                "categoriesEn": split_list(entry.values.get("categoriesEn", "")),
                "muscles": split_list(entry.values.get("muscles", "")),
                "musclesEn": split_list(entry.values.get("musclesEn", "")),
                "musclesLatin": split_list(entry.values.get("musclesLatin", "")),
                "shortDescription": clean_text(entry.values.get("shortDescription", "")),
                "shortDescriptionEn": clean_text(entry.values.get("shortDescriptionEn", "")),
                "notes": clean_text(entry.values.get("notes", "")),
                "notesEn": clean_text(entry.values.get("notesEn", "")),
            }
            item["i18n"] = {
                "name": make_i18n(item["name"], item["nameEn"]),
                "categories": make_i18n_list(entry.values.get("categories", ""), entry.values.get("categoriesEn", "")),
                "muscles": make_i18n_list(entry.values.get("muscles", ""), entry.values.get("musclesEn", "")),
                "shortDescription": make_i18n(item["shortDescription"], item["shortDescriptionEn"]),
                "notes": make_i18n(item["notes"], item["notesEn"]),
            }
            exercises.append(item)

            if self.animations_dir and not self.animation_path_for(entry.index).exists():
                missing_animations.append(f"{entry.index}.json")

        exercises.sort(key=lambda item: item["index"])
        output_json_path.parent.mkdir(parents=True, exist_ok=True)
        output_json_path.write_text(json.dumps(exercises, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        duplicate_names = [entry for entry in self.entries if entry.duplicate_name_count > 1]
        report_lines = [
            "Exercise Audit Tool Export",
            f"Zeit: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Excel: {self.excel_path}",
            f"Sheet: {self.sheet_name}",
            f"Header-Zeile: {self.header_row}",
            f"Exportierte Übungen: {len(exercises)}",
            f"Dubletten nach Name: {len(duplicate_names)}",
            f"Fehlende Animationen: {len(missing_animations)}",
            "",
            "Warnungen:",
        ]
        report_lines.extend(warnings or ["Keine."])
        report_lines.append("")
        report_lines.append("Fehlende Animationen:")
        report_lines.extend(missing_animations or ["Keine."])

        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text("\n".join(report_lines) + "\n", encoding="utf-8")

        return {
            "exported_count": len(exercises),
            "warnings": warnings,
            "missing_animations": missing_animations,
            "report_path": str(report_path),
            "output_json_path": str(output_json_path),
        }

    def table_rows(self, filter_mode: str, search_text: str) -> list[ExerciseEntry]:
        text = normalize_name_key(search_text)
        rows: list[ExerciseEntry] = []
        for entry in self.entries:
            if filter_mode == "gender_variant_candidates":
                continue
            if filter_mode == "duplicates" and entry.duplicate_name_count <= 1:
                continue
            if filter_mode == "suspicious" and not entry.suspicious_name:
                continue
            if filter_mode == "json_mismatch" and not entry.json_mismatch:
                continue
            if filter_mode == "equipment_conflict" and not entry.equipment_conflict:
                continue
            haystack = " | ".join(
                [
                    str(entry.index),
                    entry.values.get("name", ""),
                    entry.values.get("nameEn", ""),
                    entry.values.get("originalName", ""),
                    entry.values.get("packName", ""),
                ]
            )
            if text and text not in normalize_name_key(haystack):
                continue
            rows.append(entry)
        return rows
